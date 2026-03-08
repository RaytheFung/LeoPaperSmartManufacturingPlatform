"""Reporting utilities for the smart manufacturing ETL pipeline."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Callable, Dict, List

import pandas as pd
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


@dataclass
class ReportContext:
    energy_aggregated: pd.DataFrame
    csi_df: pd.DataFrame
    mes_df: pd.DataFrame
    mapping_result: Dict[str, object]
    partial_matches: Dict[str, List]


class ETLReporter:
    def __init__(self, context: ReportContext):
        self.energy_aggregated = context.energy_aggregated
        self.csi_df = context.csi_df
        self.mes_df = context.mes_df
        self.mapping_result = context.mapping_result
        self.partial_matches = context.partial_matches
        self.integrated_metrics: pd.DataFrame | None = None

    # ------------------------------------------------------------------
    # Metric calculations
    # ------------------------------------------------------------------
    def calculate_integrated_metrics(self) -> pd.DataFrame:
        print("\n=== CALCULATING INTEGRATED METRICS ===")
        three_way_matches = self.mapping_result['three_way_matches']
        integrated_data: List[Dict] = []

        for match in three_way_matches:
            machine_id = match['machine_id']
            energy_info = self.energy_aggregated.loc[machine_id]
            csi_subset = self.csi_df[self.csi_df['機台編號'] == match['csi']]
            total_good = csi_subset['正品數量'].sum() if '正品數量' in csi_subset.columns else 0
            total_bad = csi_subset['廢品數量'].sum() if '廢品數量' in csi_subset.columns else 0
            mes_subset = self.mes_df[self.mes_df['資源'] == match['mes']]
            total_planned = mes_subset['計劃生產數量'].sum() if '計劃生產數量' in mes_subset.columns else 0
            total_actual = mes_subset['實際完成數量'].sum() if '實際完成數量' in mes_subset.columns else 0

            defect_rate = (total_bad / (total_good + total_bad) * 100) if (total_good + total_bad) > 0 else 0
            plan_achievement = (total_actual / total_planned * 100) if total_planned > 0 else 0
            energy_per_unit = energy_info['total_kwh'] / total_good if total_good > 0 else 0

            integrated_data.append({
                'machine_id': machine_id,
                'energy_id': machine_id,
                'csi_id': match['csi'],
                'mes_id': match['mes'],
                'total_kwh': energy_info['total_kwh'],
                'good_products': total_good,
                'defect_products': total_bad,
                'defect_rate': defect_rate,
                'planned_quantity': total_planned,
                'actual_quantity': total_actual,
                'plan_achievement': plan_achievement,
                'kwh_per_unit': energy_per_unit,
                'components': energy_info['unique_components']
            })

        self.integrated_metrics = pd.DataFrame(integrated_data)
        print(f"\nIntegrated metrics calculated for {len(integrated_data)} machines")

        if len(integrated_data) > 0:
            print("\nKey insights:")
            print(f"  Average defect rate: {self.integrated_metrics['defect_rate'].mean():.4f}%")
            print(f"  Average plan achievement: {self.integrated_metrics['plan_achievement'].mean():.1f}%")
            print(f"  Average energy per good unit: {self.integrated_metrics['kwh_per_unit'].mean():.4f} kWh")

        return self.integrated_metrics

    # ------------------------------------------------------------------
    # Report generation
    # ------------------------------------------------------------------
    def generate_report(self, filename: str) -> None:
        print(f"\n=== GENERATING ENHANCED REPORT: {filename} ===")
        if self.integrated_metrics is None:
            self.calculate_integrated_metrics()

        wb = Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, header_fill, header_font)

        ws_three_way = wb.create_sheet("Three-Way Matches")
        self._create_three_way_sheet(ws_three_way, header_fill, header_font)

        ws_metrics = wb.create_sheet("Integrated Metrics")
        self._create_metrics_sheet(ws_metrics, header_fill, header_font)

        ws_partial = wb.create_sheet("Partial Matches")
        self._create_partial_matches_sheet(ws_partial, header_fill, header_font, subheader_fill)

        wb.save(filename)
        print(f"✅ Enhanced report saved to {filename}")

    # ------------------------------------------------------------------
    # Sheet helpers (ported from original implementation)
    # ------------------------------------------------------------------
    @staticmethod
    def _append_styled_row(ws, values, fill=None, font=None, alignment: Alignment | None = None):
        ws.append(values)
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment

    @staticmethod
    def _append_section_header(ws, title: str, fill: PatternFill | None = None):
        ws.append([title])
        row_idx = ws.max_row
        if fill:
            ws.cell(row=row_idx, column=1).fill = fill

    def _append_partial_section(
        self,
        ws,
        title: str,
        headers: List[str],
        records: List[Dict],
        row_builder: Callable[[Dict], List],
        header_fill: PatternFill,
        header_font: Font,
        subheader_fill: PatternFill,
    ):
        self._append_section_header(ws, title, fill=subheader_fill)
        self._append_styled_row(ws, headers, fill=header_fill, font=header_font)
        for record in records:
            ws.append(row_builder(record))

    def _create_summary_sheet(self, ws, header_fill, header_font):
        ws.append(["Smart Manufacturing ETL Summary"])
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

        ws.append([])
        ws.append(["Metric", "Value"])
        ws["A3"].fill = header_fill
        ws["B3"].fill = header_fill
        ws["A3"].font = header_font
        ws["B3"].font = header_font

        stats = self.mapping_result['mapping_stats']
        rows = [
            ["Energy original rows", stats['energy_original_rows']],
            ["Energy unique machines", stats['energy_unique_machines']],
            ["CSI machines", stats['csi_machines']],
            ["MES machines", stats['mes_machines']],
            ["Three-way matches", stats['three_way_matches']],
            ["MES coverage", stats['mes_coverage_percent']],
        ]

        for row in rows:
            ws.append(row)

    def _create_three_way_sheet(self, ws, header_fill, header_font):
        self._append_styled_row(
            ws,
            ["Machine ID", "CSI ID", "MES ID", "Total kWh", "Pattern", "Components", "Energy Samples"],
            fill=header_fill,
            font=header_font,
        )

        for match in self.mapping_result['three_way_matches']:
            ws.append([
                match['machine_id'],
                match['csi'],
                match['mes'],
                match['total_kwh'],
                match['pattern'],
                match['components'],
                ", ".join(match['energy_samples'])
            ])

    def _create_metrics_sheet(self, ws, header_fill, header_font):
        self._append_styled_row(ws, list(self.integrated_metrics.columns), fill=header_fill, font=header_font)

        for _, row in self.integrated_metrics.iterrows():
            ws.append(list(row.values))

    def _create_partial_matches_sheet(self, ws, header_fill, header_font, subheader_fill):
        self._append_styled_row(ws, ["Partial Match Category", "Count", "Description"], fill=header_fill, font=header_font)

        partial = self.partial_matches
        summary_rows = [
            ['Energy-CSI only', len(partial['energy_csi_only']), 'Machines in Energy & CSI but not MES'],
            ['Energy-MES only', len(partial['energy_mes_only']), 'Machines in Energy & MES but not CSI'],
            ['CSI-MES only', len(partial['csi_mes_only']), 'Machines in CSI & MES but not Energy'],
            ['Energy only', len(partial['energy_only']), 'Machines only present in Energy'],
            ['CSI only', len(partial['csi_only']), 'Machines only present in CSI'],
            ['MES only', len(partial['mes_only']), 'Machines only present in MES'],
        ]
        for row in summary_rows:
            ws.append(row)

        ws.append([])
        self._append_partial_section(
            ws,
            "Energy-CSI only",
            ["Energy", "CSI", "Top Samples", "Total kWh"],
            partial['energy_csi_only'],
            lambda item: [
                item['energy'],
                item['csi'],
                ", ".join(item['energy_samples']),
                item['total_kwh'],
            ],
            header_fill,
            header_font,
            subheader_fill,
        )

        ws.append([])
        self._append_partial_section(
            ws,
            "Energy-MES only",
            ["Energy", "MES", "Top Samples", "Total kWh"],
            partial['energy_mes_only'],
            lambda item: [
                item['energy'],
                item['mes'],
                ", ".join(item['energy_samples']),
                item['total_kwh'],
            ],
            header_fill,
            header_font,
            subheader_fill,
        )

        ws.append([])
        self._append_partial_section(
            ws,
            "CSI-MES only",
            ["CSI", "MES"],
            partial['csi_mes_only'],
            lambda item: [item['csi'], item['mes']],
            header_fill,
            header_font,
            subheader_fill,
        )
