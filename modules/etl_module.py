"""
Module 1: ETL Pipeline with File Upload
Allows users to upload monthly data files and process machine mappings
"""

import streamlit as st
import pandas as pd
import sqlite3
import json
import re
from collections import Counter
from datetime import date, datetime
from io import BytesIO
import os
from pathlib import Path

from core.bronze_raw_store import BronzeRawStore
from core.canonical_materializer import CanonicalMaterializer
from core.enhanced_etl_solution_CURRENT import EnhancedSmartManufacturingETL
from core.runtime_capabilities import suppress_write_controls
from core.runtime_mode import normalize_runtime_mode
from core.runtime_paths import (
    get_csi_dataset_dir,
    get_data_dir,
    get_database_path,
    get_energy_dataset_dir,
    get_extended_raw_dataset_root,
    get_mes_dataset_dir,
    get_raw_dataset_root,
    get_repo_root,
)
from core.source_manifest_discovery import resolve_manifest_month_sources


MONTH_NAME_OPTIONS = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]

MONTH_ALIAS_TO_NAME = {
    "jan": "January",
    "january": "January",
    "feb": "February",
    "february": "February",
    "mar": "March",
    "march": "March",
    "apr": "April",
    "april": "April",
    "may": "May",
    "jun": "June",
    "june": "June",
    "jul": "July",
    "july": "July",
    "aug": "August",
    "august": "August",
    "sep": "September",
    "sept": "September",
    "september": "September",
    "oct": "October",
    "october": "October",
    "nov": "November",
    "november": "November",
    "dec": "December",
    "december": "December",
}

MONTH_TOKEN_PATTERN = re.compile(
    r"(?<![a-z])("
    + "|".join(sorted(MONTH_ALIAS_TO_NAME.keys(), key=len, reverse=True))
    + r")(?![a-z])",
    re.IGNORECASE,
)
YEAR_TOKEN_PATTERN = re.compile(r"(?<!\d)(20\d{2})(?!\d)")
TEXTUAL_MONTH_YEAR_PATTERN = re.compile(
    r"("
    + "|".join(sorted(MONTH_ALIAS_TO_NAME.keys(), key=len, reverse=True))
    + r")[\s_\-/.,]*(20\d{2})",
    re.IGNORECASE,
)
NUMERIC_MONTH_YEAR_PATTERN = re.compile(
    r"(20\d{2})\s*[-/_.年]\s*(1[0-2]|0?[1-9])(?:\s*[月/_.-])?",
    re.IGNORECASE,
)
CHINESE_MONTH_PATTERN = re.compile(r"(?<!\d)(1[0-2]|0?[1-9])\s*月")

EXTENSION_MONTH_SOURCE_MAPPINGS = {
    "July 2025": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表__2025.7.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2025年7月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "complete", "csi": "complete", "mes": "complete"},
        "notes": [
            "July 2025 CSI is accepted after direct file audit; the earlier Aug-start wording was inconsistent with the actual source package.",
        ],
    },
    "August 2025": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2025.8-10.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2025年8月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "partial", "csi": "complete", "mes": "complete"},
        "notes": [
            "Energy remains usable but excludes the 2025-08-17 08:00-17:00 sentinel anomaly rows for 1024-10032/024-147印刷機UV.",
        ],
    },
    "September 2025": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2025.8-10.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2025年9月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "complete", "csi": "complete", "mes": "complete"},
        "notes": [],
    },
    "October 2025": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2025.8-10.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2025年10月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "partial", "csi": "complete", "mes": "complete"},
        "notes": [
            "Energy contains localized partial meter-month cases in October 2025 and is accepted with explicit flags rather than a global block.",
        ],
    },
    "November 2025": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2025.11-12.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2025年11月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "partial", "csi": "complete", "mes": "complete"},
        "notes": [
            "Energy contains a localized November 2025 partial meter-month case for 印刷機1024-10009（IR+UV）.",
        ],
    },
    "December 2025": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2025.11-12.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2025年12月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "complete", "csi": "complete", "mes": "complete"},
        "notes": [],
    },
    "January 2026": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2026.1-3.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2026年1月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "partial", "csi": "complete", "mes": "complete"},
        "notes": [
            "Energy contains localized January 2026 partial meter-month cases and is accepted with flags.",
        ],
    },
    "February 2026": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2026.1-3.xlsx"],
        "csi_file": "CSI(July2025 to Feb2026)/CSI印刷心電圖報表2026年2月.xls",
        "mes_file": "印刷機MES生產數據-2025年3月1日至2026年2月28日.xlsx",
        "family_status": {"energy": "partial", "csi": "partial", "mes": "partial"},
        "notes": [
            "Energy contains localized February 2026 partial meter-month cases and is accepted with flags.",
            "CSI/MES rows for unresolved machine 1262-00012 remain quarantined because no safe canonical mapping was proven in Task13.",
        ],
    },
    "March 2026": {
        "energy_files": ["Energy(July2025-March2026)/能耗、費用報表_2026.1-3.xlsx"],
        "csi_file": None,
        "mes_file": None,
        "family_status": {"energy": "blocked", "csi": "blocked", "mes": "blocked"},
        "notes": [
            "March 2026 is intentionally excluded from Task13 scope even though grouped energy files include March rows and MES status/create timestamps extend into early March.",
        ],
    },
}


def _month_bounds(month_year: str) -> tuple[pd.Timestamp, pd.Timestamp]:
    parsed = datetime.strptime(month_year.strip(), "%B %Y")
    start = pd.Timestamp(year=parsed.year, month=parsed.month, day=1)
    if parsed.month == 12:
        end = pd.Timestamp(year=parsed.year + 1, month=1, day=1)
    else:
        end = pd.Timestamp(year=parsed.year, month=parsed.month + 1, day=1)
    return start, end


def _series_in_target_month(series: pd.Series, month_year: str) -> pd.Series:
    start, end = _month_bounds(month_year)
    parsed = pd.to_datetime(series, errors="coerce")
    return (parsed >= start) & (parsed < end)


def _scope_source_dataframe_to_month(df: pd.DataFrame | None, source_kind: str, month_year: str) -> pd.DataFrame | None:
    if df is None:
        return None
    if df.empty:
        return df.copy()

    if source_kind == "energy":
        if "datetime" not in df.columns:
            return df.iloc[0:0].copy()
        mask = _series_in_target_month(df["datetime"], month_year)
    elif source_kind == "csi":
        mask = pd.Series(False, index=df.index)
        for column_name in ("工程開始時間", "工程結束時間", "準備結束時間", "班次內日期"):
            if column_name in df.columns:
                mask = mask | _series_in_target_month(df[column_name], month_year)
    elif source_kind == "mes":
        if "報工時間" not in df.columns:
            return df.iloc[0:0].copy()
        mask = _series_in_target_month(df["報工時間"], month_year)
    else:
        return df.copy()

    return df.loc[mask].copy()


def _scope_etl_state_to_month(etl, month_year: str) -> None:
    if not hasattr(etl, "state"):
        return
    etl.state.energy_data = _scope_source_dataframe_to_month(etl.state.energy_data, "energy", month_year)
    etl.state.csi_data = _scope_source_dataframe_to_month(etl.state.csi_data, "csi", month_year)
    etl.state.mes_data = _scope_source_dataframe_to_month(etl.state.mes_data, "mes", month_year)
    etl.state.energy_aggregated = None
    etl.state.machine_mapping = None
    etl.state.partial_matches = None
    etl.state.integrated_metrics = None


def _uploaded_file_suffix(uploaded_file, default_suffix: str = ".xlsx") -> str:
    filename = str(getattr(uploaded_file, "name", "") or "").strip()
    suffix = Path(filename).suffix.lower()
    return suffix if suffix in {".xls", ".xlsx", ".xlsm"} else default_suffix


def _resolve_default_data_root_for_month(month_year: str) -> Path:
    if month_year in EXTENSION_MONTH_SOURCE_MAPPINGS:
        return get_extended_raw_dataset_root()
    return get_raw_dataset_root()


def _resolve_extension_source_mapping(month_year: str, data_root: Path | str | None = None) -> dict[str, object]:
    spec = EXTENSION_MONTH_SOURCE_MAPPINGS[month_year]
    dataset_root = Path(data_root) if data_root is not None else get_extended_raw_dataset_root()
    resolved = {
        "dataset_root": str(dataset_root),
        "energy_files": [str(dataset_root / relative_path) for relative_path in spec["energy_files"]],
        "csi_file": str(dataset_root / spec["csi_file"]) if spec.get("csi_file") else None,
        "mes_file": str(dataset_root / spec["mes_file"]) if spec.get("mes_file") else None,
        "family_status": dict(spec["family_status"]),
        "notes": list(spec["notes"]),
    }
    family_status = resolved["family_status"]
    if all(status == "complete" for status in family_status.values()):
        resolved["backfill_readiness"] = "ready"
    elif any(status == "blocked" for status in family_status.values()):
        resolved["backfill_readiness"] = "blocked"
    else:
        resolved["backfill_readiness"] = "ready_with_flags"
    return resolved


def _build_extension_source_availability_dataframe(data_root: Path | str | None = None) -> pd.DataFrame:
    rows = []
    for month_year in EXTENSION_MONTH_SOURCE_MAPPINGS:
        mapping = _resolve_extension_source_mapping(month_year, data_root=data_root)
        file_candidates = [
            *mapping["energy_files"],
            *( [mapping["csi_file"]] if mapping.get("csi_file") else [] ),
            *( [mapping["mes_file"]] if mapping.get("mes_file") else [] ),
        ]
        missing_files = [path for path in file_candidates if path and not Path(path).exists()]
        rows.append(
            {
                "Month": month_year,
                "Energy": mapping["family_status"]["energy"].title(),
                "CSI": mapping["family_status"]["csi"].title(),
                "MES": mapping["family_status"]["mes"].title(),
                "Backfill Readiness": (
                    "Blocked"
                    if missing_files or mapping["backfill_readiness"] == "blocked"
                    else "Ready with Flags"
                    if mapping["backfill_readiness"] == "ready_with_flags"
                    else "Ready"
                ),
                "Notes": " | ".join(mapping["notes"]) or "none",
                "Missing Files": " | ".join(missing_files) or "none",
            }
        )
    return pd.DataFrame(rows)


def _compare_resolved_source_payloads(
    legacy_payload: dict[str, object],
    manifest_payload: dict[str, object],
) -> dict[str, object]:
    compared_fields = (
        "energy_files",
        "csi_file",
        "mes_file",
        "family_status",
        "backfill_readiness",
    )
    differences = []
    for field_name in compared_fields:
        legacy_value = legacy_payload.get(field_name)
        manifest_value = manifest_payload.get(field_name)
        if legacy_value != manifest_value:
            differences.append(
                {
                    "field": field_name,
                    "legacy": legacy_value,
                    "manifest": manifest_value,
                }
            )
    return {"matches": not differences, "differences": differences}


def _source_discovery_error_payload(mode: str, exc: Exception) -> dict[str, object]:
    return {
        "mode": mode,
        "error_type": exc.__class__.__name__,
        "message": str(exc),
        "blocked": "blocked" in str(exc).lower(),
    }


class ETLPipelineModule:
    HISTORICAL_MONTH_FILE_MAPPINGS = {
        "January": {
            "energy": [
                "能耗、費用報表Jan(1-10).xlsx",
                "能耗、費用報表Jan(11-21).xlsx",
                "能耗、費用報表Jan(22-31).xlsx",
            ],
            "csi": "CSI印刷心電圖報表Jan.xlsx",
            "mes": "MES生產數據Jan(Printer).xlsx",
        },
        "February": {
            "energy": [
                "能耗、費用報表Feb(1-10).xlsx",
                "能耗、費用報表Feb(11-21).xlsx",
                "能耗、費用報表Feb(22-28).xlsx",
            ],
            "csi": "CSI印刷心電圖報表Feb.xlsx",
            "mes": "MES生產數據Feb(Printer).xlsx",
        },
        "March": {
            "energy": [
                "能耗、費用報表March(1-10).xlsx",
                "能耗、費用報表March(11-21).xlsx",
                "能耗、費用報表March(22-31).xlsx",
            ],
            "csi": "CSI印刷心電圖報表March.xlsx",
            "mes": "MES生產數據March(Printer).xlsx",
        },
        "April": {
            "energy": [
                "能耗、費用報表April(1-10).xlsx",
                "能耗、費用報表April(11-21).xlsx",
                "能耗、費用報表April(22-30).xlsx",
            ],
            "csi": "CSI印刷心電圖報表April.xlsx",
            "mes": "MES生產數據April(Printer).xlsx",
        },
        "May": {
            "energy": [
                "能耗、費用報表May(1-31).xlsx",
            ],
            "csi": "CSI印刷心電圖報表May.xlsx",
            "mes": "MES生產數據May(Printer).xlsx",
        },
        "June": {
            "energy": [
                "能耗、費用報表June(1-30).xlsx",
            ],
            "csi": "CSI印刷心電圖報表June.xlsx",
            "mes": "MES生產數據June(Printer).xlsx",
        },
    }
    MAINTENANCE_FILE_CANDIDATES = (
        "印刷機維修記錄清單（2025年全年）.xlsx",
        "Maintenance RecordJan to Jul.xlsx",
    )

    def __init__(self, db_path=None, *, initialize_schema: bool = True):
        self.db_path = str(db_path or get_database_path())
        self.bronze_store = BronzeRawStore(self.db_path) if initialize_schema else None
        if initialize_schema:
            self.init_database()
        
    def init_database(self):
        """Initialize SQLite database with required tables"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Create tables if they don't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS etl_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_date TIMESTAMP,
                month_processed TEXT,
                energy_files_count INTEGER,
                three_way_matches INTEGER,
                match_rate REAL,
                status TEXT,
                details TEXT,
                display_order INTEGER
            )
        ''')
        
        # Add display_order column if it doesn't exist (for existing databases)
        cursor.execute("PRAGMA table_info(etl_runs)")
        columns = [column[1] for column in cursor.fetchall()]
        if 'display_order' not in columns:
            cursor.execute('ALTER TABLE etl_runs ADD COLUMN display_order INTEGER')
            # Set initial display order based on run_date
            cursor.execute('''
                UPDATE etl_runs 
                SET display_order = (
                    SELECT COUNT(*) 
                    FROM etl_runs AS e2 
                    WHERE e2.run_date <= etl_runs.run_date
                )
            ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS machine_inventory (
                machine_id TEXT,
                system_type TEXT,
                first_seen_date TEXT,
                last_seen_date TEXT,
                is_active INTEGER,
                PRIMARY KEY (machine_id, system_type)
            )
        ''')
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS three_way_matches (
                machine_id TEXT PRIMARY KEY,
                energy_pattern TEXT,
                csi_id TEXT,
                mes_id TEXT,
                first_matched_date TEXT,
                last_confirmed_date TEXT
            )
        ''')
        
        # Create monthly presence tracking table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS machine_monthly_presence (
                machine_id TEXT,
                month_year TEXT,
                system_type TEXT,
                is_three_way_match INTEGER DEFAULT 0,
                PRIMARY KEY (machine_id, month_year, system_type)
            )
        ''')
        
        conn.commit()
        conn.close()

    def get_historical_month_file_mappings(self, data_root=None):
        dataset_root = Path(data_root) if data_root is not None else get_raw_dataset_root()
        energy_dir = get_energy_dataset_dir(dataset_root)
        csi_dir = get_csi_dataset_dir(dataset_root)
        mes_dir = get_mes_dataset_dir(dataset_root)

        resolved = {}
        for month_name, file_mapping in self.HISTORICAL_MONTH_FILE_MAPPINGS.items():
            resolved[f"{month_name} 2025"] = {
                "dataset_root": str(dataset_root),
                "energy_files": [str(energy_dir / file_name) for file_name in file_mapping["energy"]],
                "csi_file": str(csi_dir / file_mapping["csi"]),
                "mes_file": str(mes_dir / file_mapping["mes"]),
                "maintenance_file": self._resolve_historical_maintenance_file(dataset_root),
            }
        for month_year in EXTENSION_MONTH_SOURCE_MAPPINGS:
            resolved[month_year] = _resolve_extension_source_mapping(month_year, data_root=data_root)
        return resolved

    def resolve_historical_month_sources(self, month_year, data_root=None, *, discovery_mode: str = "legacy"):
        month_key = month_year.strip()
        normalized_mode = str(discovery_mode or "legacy").strip().lower()
        if normalized_mode not in {"legacy", "manifest", "compare"}:
            raise ValueError(
                "Unsupported source discovery mode "
                f"{discovery_mode!r}; expected legacy, manifest, or compare."
            )

        if normalized_mode == "manifest":
            manifest_data_root = data_root if data_root is not None else _resolve_default_data_root_for_month(month_key)
            source_files = resolve_manifest_month_sources(
                month_key,
                data_root=manifest_data_root,
            )
            source_files["source_discovery_mode"] = "manifest"
            return source_files

        if normalized_mode == "compare":
            return self._resolve_historical_month_sources_compare(month_key, data_root=data_root)

        return self._resolve_historical_month_sources_legacy(month_key, data_root=data_root)

    def _resolve_historical_month_sources_legacy(self, month_key, data_root=None):
        historical_mappings = self.get_historical_month_file_mappings(data_root)
        if month_key not in historical_mappings:
            raise ValueError(f"No historical source mapping is defined for {month_key}.")

        source_files = historical_mappings[month_key]
        if source_files.get("backfill_readiness") == "blocked":
            notes = source_files.get("notes") or ["The requested month is explicitly blocked for controlled backfill."]
            raise ValueError(
                f"Historical backfill is blocked for {month_key}: {' '.join(notes)}"
            )
        missing_files = [
            file_path
            for file_path in [
                *source_files["energy_files"],
                *( [source_files["csi_file"]] if source_files.get("csi_file") else [] ),
                *( [source_files["mes_file"]] if source_files.get("mes_file") else [] ),
            ]
            if not Path(file_path).exists()
        ]
        if missing_files:
            raise ValueError(
                "Historical backfill cannot run because source files are missing for "
                f"{month_key}: {', '.join(missing_files)}"
            )

        return source_files

    def _resolve_historical_month_sources_compare(self, month_key, data_root=None):
        legacy_payload = None
        manifest_payload = None
        legacy_error = None
        manifest_error = None

        try:
            legacy_payload = self._resolve_historical_month_sources_legacy(month_key, data_root=data_root)
        except Exception as exc:
            legacy_error = _source_discovery_error_payload("legacy", exc)

        try:
            manifest_data_root = data_root if data_root is not None else _resolve_default_data_root_for_month(month_key)
            manifest_payload = resolve_manifest_month_sources(month_key, data_root=manifest_data_root)
        except Exception as exc:
            manifest_error = _source_discovery_error_payload("manifest", exc)

        if legacy_payload is not None and manifest_payload is not None:
            comparison = _compare_resolved_source_payloads(legacy_payload, manifest_payload)
            result = dict(legacy_payload)
            result["source_discovery_mode"] = "compare"
            result["manifest_equivalence"] = comparison
            return result

        equivalence = {
            "matches": False,
            "differences": [],
            "legacy_error": legacy_error,
            "manifest_error": manifest_error,
        }
        if legacy_error and manifest_error:
            equivalence["both_blocked"] = bool(legacy_error.get("blocked") and manifest_error.get("blocked"))

        if legacy_payload is not None:
            result = dict(legacy_payload)
            result["source_discovery_mode"] = "compare"
            result["manifest_equivalence"] = equivalence
            return result

        return {
            "dataset_root": str(data_root) if data_root is not None else str(_resolve_default_data_root_for_month(month_key)),
            "energy_files": [],
            "csi_file": None,
            "mes_file": None,
            "family_status": {},
            "notes": [],
            "backfill_readiness": "blocked" if (legacy_error or manifest_error) else "unknown",
            "source_discovery_mode": "compare",
            "manifest_equivalence": equivalence,
        }

    def build_extension_source_availability(self, data_root=None):
        return _build_extension_source_availability_dataframe(data_root=data_root)

    def run_historical_canonical_backfill(self, month_years, data_root=None):
        requested_months = [month for month in month_years if str(month).strip()]
        if not requested_months:
            return {
                "status": "error",
                "requested_months": [],
                "monthly_results": [],
                "failed_months": [],
                "legacy_unified_view_bypassed": True,
                "message": "No historical months were requested.",
            }

        materializer = CanonicalMaterializer(self.db_path)
        monthly_results = []
        failed_months = []

        for month_year in requested_months:
            try:
                source_files = self.resolve_historical_month_sources(month_year, data_root=data_root)
                etl = EnhancedSmartManufacturingETL()
                etl.extract_all_sources(
                    source_files["energy_files"],
                    source_files["csi_file"],
                    source_files["mes_file"],
                )
                _scope_etl_state_to_month(etl, month_year)
                mapping_results = etl.create_comprehensive_mapping()
                self.save_etl_results(mapping_results, month_year, etl)
                canonical_result = materializer.materialize_backfill_month(month_year)
                monthly_results.append(
                    {
                        "target_month": month_year,
                        "source_files": source_files,
                        "source_notes": list(source_files.get("notes") or []),
                        "source_readiness": source_files.get("backfill_readiness"),
                        "canonical_materialization": canonical_result,
                    }
                )
            except Exception as exc:
                failed_months.append(
                    {
                        "target_month": month_year,
                        "message": str(exc),
                    }
                )

        if failed_months and monthly_results:
            status = "partial_error"
        elif failed_months:
            status = "error"
        else:
            status = "success"

        return {
            "status": status,
            "requested_months": requested_months,
            "successful_months": [result["target_month"] for result in monthly_results],
            "failed_months": failed_months,
            "monthly_results": monthly_results,
            "legacy_unified_view_bypassed": True,
        }

    def _resolve_historical_maintenance_file(self, data_root):
        dataset_root = Path(data_root)
        maintenance_dirs = [
            dataset_root / "(12:3:2026) Maintenance",
            dataset_root,
        ]
        for parent in maintenance_dirs:
            for file_name in self.MAINTENANCE_FILE_CANDIDATES:
                candidate = parent / file_name
                if candidate.exists():
                    return str(candidate)
        return None

    def save_etl_results(self, mapping_results, month_name, etl_instance=None):
        """Save ETL results to database including actual extracted data"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        def ensure_table_columns(table_name, columns):
            cursor.execute(f"PRAGMA table_info({table_name})")
            existing_columns = {column[1] for column in cursor.fetchall()}
            for column_name, column_type in columns.items():
                if column_name not in existing_columns:
                    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}")
        
        # Create tables for storing actual ETL data if they don't exist
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS etl_energy_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_year TEXT,
                pattern TEXT,
                datetime TIMESTAMP,
                electricity_kwh REAL,
                machine_components TEXT,
                canonical_machine_id TEXT,
                matched_on TEXT,
                matched_value TEXT,
                exception_applied INTEGER,
                source_system TEXT,
                scope_status TEXT,
                join_status TEXT
            )
        ''')
        ensure_table_columns('etl_energy_data', {
            'canonical_machine_id': 'TEXT',
            'matched_on': 'TEXT',
            'matched_value': 'TEXT',
            'exception_applied': 'INTEGER',
            'source_system': 'TEXT',
            'scope_status': 'TEXT',
            'join_status': 'TEXT',
        })
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS etl_csi_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_year TEXT,
                machine_id TEXT,
                start_time TIMESTAMP,
                end_time TIMESTAMP,
                setup_start TIMESTAMP,
                setup_end TIMESTAMP,
                material TEXT,
                order_id TEXT,
                good_qty REAL,
                efficiency REAL,
                actual_speed REAL,
                team_leader TEXT,
                team_member_1 TEXT,
                team_member_2 TEXT,
                team_member_3 TEXT,
                team_member_4 TEXT,
                canonical_machine_id TEXT,
                matched_on TEXT,
                matched_value TEXT,
                exception_applied INTEGER,
                source_system TEXT,
                scope_status TEXT,
                join_status TEXT
            )
        ''')
        ensure_table_columns('etl_csi_data', {
            'canonical_machine_id': 'TEXT',
            'matched_on': 'TEXT',
            'matched_value': 'TEXT',
            'exception_applied': 'INTEGER',
            'source_system': 'TEXT',
            'scope_status': 'TEXT',
            'join_status': 'TEXT',
        })
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS etl_mes_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                month_year TEXT,
                resource TEXT,
                task TEXT,
                order_number TEXT,
                material_code TEXT,
                planned_qty REAL,
                planned_start TIMESTAMP,
                planned_end TIMESTAMP,
                canonical_machine_id TEXT,
                matched_on TEXT,
                matched_value TEXT,
                exception_applied INTEGER,
                source_system TEXT,
                scope_status TEXT,
                join_status TEXT
            )
        ''')
        ensure_table_columns('etl_mes_data', {
            'canonical_machine_id': 'TEXT',
            'matched_on': 'TEXT',
            'matched_value': 'TEXT',
            'exception_applied': 'INTEGER',
            'source_system': 'TEXT',
            'scope_status': 'TEXT',
            'join_status': 'TEXT',
        })
        
        # Store the actual data if ETL instance is provided
        if etl_instance:
            energy_data = _scope_source_dataframe_to_month(
                getattr(etl_instance, "energy_data", None),
                "energy",
                month_name,
            )
            csi_data = _scope_source_dataframe_to_month(
                getattr(etl_instance, "csi_data", None),
                "csi",
                month_name,
            )
            mes_data = _scope_source_dataframe_to_month(
                getattr(etl_instance, "mes_data", None),
                "mes",
                month_name,
            )

            if energy_data is not None:
                self.bronze_store.write_energy_rows(energy_data)
            if csi_data is not None:
                self.bronze_store.write_csi_rows(csi_data)
            if mes_data is not None:
                self.bronze_store.write_mes_rows(mes_data)

            # Clear existing data for this month
            cursor.execute("DELETE FROM etl_energy_data WHERE month_year = ?", (month_name,))
            cursor.execute("DELETE FROM etl_csi_data WHERE month_year = ?", (month_name,))
            cursor.execute("DELETE FROM etl_mes_data WHERE month_year = ?", (month_name,))
            
            # Store energy data - use original energy_data with machine mappings
            if energy_data is not None:
                # Create a mapping from original machine names to patterns
                machine_to_pattern = {}
                if hasattr(etl_instance, 'energy_aggregated'):
                    for machine_id, row in etl_instance.energy_aggregated.iterrows():
                        # Map each original name to the machine_id (which is the pattern)
                        for orig_name in row.get('original_names', []):
                            machine_to_pattern[orig_name] = machine_id
                
                # Save the detailed energy data with patterns
                for _, row in energy_data.iterrows():
                    machine_name = row.get('machine', '')
                    pattern = row.get('canonical_machine_id') or machine_to_pattern.get(machine_name, '')
                    
                    # Skip if no pattern found (non-matched machines)
                    if not pattern:
                        continue
                        
                    cursor.execute('''
                        INSERT INTO etl_energy_data 
                        (month_year, pattern, datetime, electricity_kwh, machine_components,
                         canonical_machine_id, matched_on, matched_value, exception_applied,
                         source_system, scope_status, join_status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        month_name,
                        pattern,  # Use the mapped pattern
                        str(row.get('datetime', '')),  # Convert to string
                        float(row.get('electricity_kwh', 0)),  # Ensure float
                        json.dumps([machine_name]),  # Store original name as component
                        row.get('canonical_machine_id'),
                        row.get('matched_on'),
                        row.get('matched_value'),
                        int(bool(row.get('exception_applied', False))),
                        row.get('source_system'),
                        row.get('scope_status'),
                        row.get('join_status'),
                    ))
            
            # Store CSI data
            if csi_data is not None:
                for _, row in csi_data.iterrows():
                    cursor.execute('''
                        INSERT INTO etl_csi_data 
                        (month_year, machine_id, start_time, end_time, setup_start, setup_end,
                         material, order_id, good_qty, efficiency, actual_speed,
                         team_leader, team_member_1, team_member_2, team_member_3, team_member_4,
                         canonical_machine_id, matched_on, matched_value, exception_applied,
                         source_system, scope_status, join_status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        month_name,
                        row.get('機台編號', ''),
                        row.get('工程開始時間'),
                        row.get('工程結束時間'),
                        row.get('準備開始時間'),
                        row.get('準備結束時間'),
                        row.get('物料', ''),
                        row.get('作业', ''),
                        row.get('正品數量', 0),
                        row.get('效率', 0),
                        row.get('實際速度_本_時', 0),
                        row.get('機長姓名1', ''),
                        row.get('機組人員姓名1', ''),
                        row.get('機組人員姓名2', ''),
                        row.get('機組人員姓名3', ''),
                        row.get('機組人員姓名4', ''),
                        row.get('canonical_machine_id'),
                        row.get('matched_on'),
                        row.get('matched_value'),
                        int(bool(row.get('exception_applied', False))),
                        row.get('source_system'),
                        row.get('scope_status'),
                        row.get('join_status'),
                    ))
            
            # Store MES data
            if mes_data is not None:
                for _, row in mes_data.iterrows():
                    cursor.execute('''
                        INSERT INTO etl_mes_data 
                        (month_year, resource, task, order_number, material_code,
                         planned_qty, planned_start, planned_end, canonical_machine_id,
                         matched_on, matched_value, exception_applied, source_system,
                         scope_status, join_status)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        month_name,
                        row.get('資源', ''),
                        row.get('任務', ''),
                        row.get('訂單號', ''),
                        row.get('物料編碼', ''),
                        row.get('計劃數量', 0),
                        row.get('計劃開始'),
                        row.get('計劃結束'),
                        row.get('canonical_machine_id'),
                        row.get('matched_on'),
                        row.get('matched_value'),
                        int(bool(row.get('exception_applied', False))),
                        row.get('source_system'),
                        row.get('scope_status'),
                        row.get('join_status'),
                    ))
        
        # Get the next display order
        cursor.execute('SELECT MAX(display_order) FROM etl_runs')
        max_order = cursor.fetchone()[0]
        next_order = (max_order or 0) + 1
        
        # Save run summary
        stats = mapping_results['mapping_stats']
        cursor.execute('''
            INSERT INTO etl_runs 
            (run_date, month_processed, energy_files_count, three_way_matches, match_rate, status, details, display_order)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            datetime.now(),
            month_name,
            stats['energy_original_rows'],
            stats['three_way_matches'],
            float(stats['mes_coverage_percent'].strip('%')),
            'Success',
            json.dumps(stats),
            next_order
        ))
        
        # First, deactivate all machines for this month's processing
        # This ensures we only show machines that are actually present in the current data
        cursor.execute('''
            UPDATE machine_inventory 
            SET is_active = 0 
            WHERE last_seen_date != ?
        ''', (month_name,))
        
        # Collect all machines seen in this processing run
        all_seen_machines = set()
        
        # Update machine inventory for three-way matches
        for match in mapping_results['three_way_matches']:
            # Update three-way matches
            cursor.execute('''
                INSERT OR REPLACE INTO three_way_matches 
                (machine_id, energy_pattern, csi_id, mes_id, first_matched_date, last_confirmed_date)
                VALUES (?, ?, ?, ?, 
                    COALESCE((SELECT first_matched_date FROM three_way_matches WHERE machine_id = ?), ?),
                    ?)
            ''', (
                match['machine_id'],
                match['machine_id'],
                match['csi'],
                match['mes'],
                match['machine_id'],
                month_name,
                month_name
            ))
            
            # Update inventory for each system
            for system, id_val in [('Energy', match['machine_id']), 
                                   ('CSI', match['csi']), 
                                   ('MES', match['mes'])]:
                all_seen_machines.add((id_val, system))
                cursor.execute('''
                    INSERT OR REPLACE INTO machine_inventory 
                    (machine_id, system_type, first_seen_date, last_seen_date, is_active)
                    VALUES (?, ?, 
                        COALESCE((SELECT first_seen_date FROM machine_inventory 
                                  WHERE machine_id = ? AND system_type = ?), ?),
                        ?, 1)
                ''', (id_val, system, id_val, system, month_name, month_name))
        
        # Also update inventory for machines in partial matches and single systems
        # to maintain accurate per-month data
        
        # Process partial matches
        if 'partial_matches' in mapping_results:
            for category, machines in mapping_results['partial_matches'].items():
                for machine_data in machines:
                    if category == 'energy_csi_only':
                        all_seen_machines.add((machine_data['machine_id'], 'Energy'))
                        all_seen_machines.add((machine_data['csi'], 'CSI'))
                    elif category == 'energy_mes_only':
                        all_seen_machines.add((machine_data['machine_id'], 'Energy'))
                        all_seen_machines.add((machine_data['mes'], 'MES'))
                    elif category == 'csi_mes_only':
                        all_seen_machines.add((machine_data['csi'], 'CSI'))
                        all_seen_machines.add((machine_data['mes'], 'MES'))
        
        # Process single system machines
        if 'single_system' in mapping_results:
            for system_key, machines in mapping_results['single_system'].items():
                system_map = {
                    'energy_only': 'Energy',
                    'csi_only': 'CSI',
                    'mes_only': 'MES'
                }
                if system_key in system_map:
                    system = system_map[system_key]
                    for machine in machines:
                        machine_id = machine if isinstance(machine, str) else str(machine)
                        all_seen_machines.add((machine_id, system))
        
        # Update all seen machines as active
        for machine_id, system in all_seen_machines:
            cursor.execute('''
                INSERT OR REPLACE INTO machine_inventory 
                (machine_id, system_type, first_seen_date, last_seen_date, is_active)
                VALUES (?, ?, 
                    COALESCE((SELECT first_seen_date FROM machine_inventory 
                              WHERE machine_id = ? AND system_type = ?), ?),
                    ?, 1)
            ''', (machine_id, system, machine_id, system, month_name, month_name))
        
        # Clear monthly presence for this month first (in case of re-processing)
        cursor.execute('DELETE FROM machine_monthly_presence WHERE month_year = ?', (month_name,))
        
        # Insert monthly presence for all seen machines
        for machine_id, system in all_seen_machines:
            # Check if this machine is in three-way matches
            is_three_way = 1 if any(m['machine_id'] == machine_id for m in mapping_results['three_way_matches']) else 0
            
            cursor.execute('''
                INSERT INTO machine_monthly_presence 
                (machine_id, month_year, system_type, is_three_way_match)
                VALUES (?, ?, ?, ?)
            ''', (machine_id, month_name, system, is_three_way))
        
        conn.commit()
        conn.close()
    
    def get_historical_summary(self, order_by='display_order'):
        """Get historical ETL run summary"""
        conn = sqlite3.connect(self.db_path)
        
        # Validate order_by parameter
        valid_orders = {
            'display_order': 'display_order ASC',
            'date_desc': 'run_date DESC',
            'date_asc': 'run_date ASC',
            'month': 'month_processed, run_date DESC',
            'match_rate': 'match_rate DESC'
        }
        order_clause = valid_orders.get(order_by, 'display_order ASC')
        
        query = f'''
            SELECT id, month_processed, run_date, three_way_matches, match_rate, display_order
            FROM etl_runs
            ORDER BY {order_clause}
        '''
        df = pd.read_sql_query(query, conn)
        conn.close()
        return df
    
    def delete_etl_run(self, run_id):
        """Delete a specific ETL run"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get the record before deletion for undo functionality
        cursor.execute('SELECT * FROM etl_runs WHERE id = ?', (run_id,))
        deleted_record = cursor.fetchone()
        
        # Delete the record
        cursor.execute('DELETE FROM etl_runs WHERE id = ?', (run_id,))
        
        # Reorder remaining records
        cursor.execute('''
            UPDATE etl_runs 
            SET display_order = (
                SELECT COUNT(*) 
                FROM etl_runs AS e2 
                WHERE e2.display_order <= etl_runs.display_order
            )
        ''')
        
        conn.commit()
        conn.close()
        return deleted_record
    
    def delete_multiple_runs(self, run_ids):
        """Delete multiple ETL runs"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Delete the records
        placeholders = ','.join('?' * len(run_ids))
        cursor.execute(f'DELETE FROM etl_runs WHERE id IN ({placeholders})', run_ids)
        
        # Reorder remaining records
        cursor.execute('''
            UPDATE etl_runs 
            SET display_order = (
                SELECT COUNT(*) 
                FROM etl_runs AS e2 
                WHERE e2.display_order <= etl_runs.display_order
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def update_display_order(self, run_id, direction):
        """Move a record up or down in display order"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Get current order
        cursor.execute('SELECT display_order FROM etl_runs WHERE id = ?', (run_id,))
        current_order = cursor.fetchone()[0]
        
        if direction == 'up' and current_order > 1:
            # Swap with previous record
            cursor.execute('''
                UPDATE etl_runs 
                SET display_order = CASE 
                    WHEN id = ? THEN display_order - 1
                    WHEN display_order = ? THEN display_order + 1
                END
                WHERE id = ? OR display_order = ?
            ''', (run_id, current_order - 1, run_id, current_order - 1))
        
        elif direction == 'down':
            # Check if not last
            cursor.execute('SELECT MAX(display_order) FROM etl_runs')
            max_order = cursor.fetchone()[0]
            
            if current_order < max_order:
                # Swap with next record
                cursor.execute('''
                    UPDATE etl_runs 
                    SET display_order = CASE 
                        WHEN id = ? THEN display_order + 1
                        WHEN display_order = ? THEN display_order - 1
                    END
                    WHERE id = ? OR display_order = ?
                ''', (run_id, current_order + 1, run_id, current_order + 1))
        
        conn.commit()
        conn.close()
    
    def reset_display_order(self):
        """Reset display order to chronological order"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE etl_runs 
            SET display_order = (
                SELECT COUNT(*) 
                FROM etl_runs AS e2 
                WHERE e2.run_date <= etl_runs.run_date
            )
        ''')
        
        conn.commit()
        conn.close()
    
    def get_machine_inventory_summary(self, active_only=True):
        """Get machine inventory summary"""
        conn = sqlite3.connect(self.db_path)
        if active_only:
            query = '''
                SELECT system_type, COUNT(DISTINCT machine_id) as machine_count
                FROM machine_inventory
                WHERE is_active = 1
                GROUP BY system_type
            '''
        else:
            query = '''
                SELECT system_type, COUNT(DISTINCT machine_id) as machine_count
                FROM machine_inventory
                GROUP BY system_type
            '''
        df = pd.read_sql_query(query, conn)
        conn.close()
        return df


def _build_upload_signature(energy_files, csi_file, mes_file):
    signature = []
    for role, uploaded_file in [
        *[("Energy", file) for file in energy_files],
        ("CSI", csi_file),
        ("MES", mes_file),
    ]:
        if uploaded_file is None:
            continue
        signature.append(
            (
                role,
                getattr(uploaded_file, "name", None),
                getattr(uploaded_file, "size", None),
            )
        )
    return tuple(signature)


def _normalize_month_name(month_value):
    if month_value is None:
        return None
    cleaned = str(month_value).strip()
    if not cleaned:
        return None
    return MONTH_ALIAS_TO_NAME.get(cleaned.lower(), cleaned if cleaned in MONTH_NAME_OPTIONS else None)


def _extract_month_year_candidates_from_text(text_value):
    if text_value is None:
        return []
    text = str(text_value).strip()
    if not text:
        return []

    candidates = []
    for month_match, year_match in TEXTUAL_MONTH_YEAR_PATTERN.findall(text):
        month_name = _normalize_month_name(month_match)
        if month_name is None:
            continue
        year = int(year_match)
        if 2020 <= year <= 2035:
            candidates.append((month_name, year))

    for year_match, month_match in NUMERIC_MONTH_YEAR_PATTERN.findall(text):
        month_num = int(month_match)
        if 1 <= month_num <= 12:
            candidates.append((MONTH_NAME_OPTIONS[month_num - 1], int(year_match)))

    return candidates


def _detect_month_year_from_filename(filename):
    normalized_name = Path(filename).name
    month_name = None
    year = None

    month_match = MONTH_TOKEN_PATTERN.search(normalized_name.lower())
    if month_match:
        month_name = _normalize_month_name(month_match.group(1))
    else:
        chinese_month_match = CHINESE_MONTH_PATTERN.search(normalized_name)
        if chinese_month_match:
            month_name = MONTH_NAME_OPTIONS[int(chinese_month_match.group(1)) - 1]

    year_match = YEAR_TOKEN_PATTERN.search(normalized_name)
    if year_match:
        year = int(year_match.group(1))

    if month_name and year is not None:
        confidence = "high"
        note = "Matched month and year in filename."
    elif month_name:
        confidence = "medium"
        note = "Matched month in filename; year still needs confirmation."
    elif year is not None:
        confidence = "low"
        note = "Matched year in filename only; month still needs confirmation."
    else:
        confidence = "none"
        note = "No reliable month/year tokens found in filename."

    return {
        "month": month_name,
        "year": year,
        "source": "filename",
        "confidence": confidence,
        "note": note,
    }


def _detect_month_year_from_excel_sample(uploaded_file):
    suffix = Path(getattr(uploaded_file, "name", "")).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        return {
            "month": None,
            "year": None,
            "source": "workbook sample",
            "confidence": "none",
            "status": "skipped",
            "note": "Workbook sample fallback is only used for .xlsx/.xlsm files.",
        }

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            filename=BytesIO(uploaded_file.getvalue()),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        return {
            "month": None,
            "year": None,
            "source": "workbook sample",
            "confidence": "none",
            "status": "error",
            "note": f"Workbook sample fallback could not open the file: {exc}",
        }

    candidates = Counter()
    try:
        for sheet_name in workbook.sheetnames[:3]:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(max_row=120, max_col=24, values_only=True):
                for value in row:
                    if isinstance(value, (datetime, pd.Timestamp)):
                        month_name = MONTH_NAME_OPTIONS[value.month - 1]
                        if 2020 <= value.year <= 2035:
                            candidates[(month_name, value.year)] += 1
                    elif isinstance(value, date):
                        month_name = MONTH_NAME_OPTIONS[value.month - 1]
                        if 2020 <= value.year <= 2035:
                            candidates[(month_name, value.year)] += 1
                    elif isinstance(value, str):
                        for candidate in _extract_month_year_candidates_from_text(value):
                            candidates[candidate] += 1
    finally:
        workbook.close()

    if not candidates:
        return {
            "month": None,
            "year": None,
            "source": "workbook sample",
            "confidence": "none",
            "status": "unresolved",
            "note": "Workbook sample did not expose one stable month/year candidate.",
        }

    if len(candidates) == 1:
        (month_name, year), count = candidates.most_common(1)[0]
        confidence = "high" if count >= 3 else "medium"
        return {
            "month": month_name,
            "year": year,
            "source": "workbook sample",
            "confidence": confidence,
            "status": "resolved",
            "note": f"Workbook sample consistently points to {month_name} {year}.",
        }

    (month_name, year), count = candidates.most_common(1)[0]
    total_hits = sum(candidates.values())
    if total_hits and count / total_hits >= 0.8 and count >= 3:
        return {
            "month": month_name,
            "year": year,
            "source": "workbook sample",
            "confidence": "medium",
            "status": "resolved",
            "note": (
                f"Workbook sample mostly points to {month_name} {year}; "
                f"{count}/{total_hits} sampled date hits agree."
            ),
        }

    candidate_text = ", ".join(
        f"{candidate_month} {candidate_year}" for (candidate_month, candidate_year), _ in candidates.most_common(3)
    )
    return {
        "month": None,
        "year": None,
        "source": "workbook sample",
        "confidence": "low",
        "status": "ambiguous",
        "note": f"Workbook sample exposed multiple month/year candidates: {candidate_text}.",
    }


def _resolve_uploaded_file_detection(uploaded_file, role_label):
    filename_detection = _detect_month_year_from_filename(uploaded_file.name)
    workbook_detection = None

    needs_workbook_fallback = filename_detection["month"] is None or filename_detection["year"] is None
    if needs_workbook_fallback:
        workbook_detection = _detect_month_year_from_excel_sample(uploaded_file)

    resolved_month = filename_detection["month"]
    resolved_year = filename_detection["year"]
    source_parts = []
    note_parts = [filename_detection["note"]]
    conflict_notes = []

    if resolved_month or resolved_year is not None:
        source_parts.append("filename")

    if workbook_detection is not None:
        note_parts.append(workbook_detection["note"])
        workbook_month = workbook_detection.get("month")
        workbook_year = workbook_detection.get("year")
        if (
            resolved_month is not None
            and workbook_month is not None
            and resolved_month != workbook_month
        ):
            conflict_notes.append(
                f"Filename month `{resolved_month}` disagrees with workbook sample `{workbook_month}`."
            )
        elif resolved_month is None and workbook_month is not None:
            resolved_month = workbook_month
            source_parts.append("workbook sample")

        if (
            resolved_year is not None
            and workbook_year is not None
            and resolved_year != workbook_year
        ):
            conflict_notes.append(
                f"Filename year `{resolved_year}` disagrees with workbook sample `{workbook_year}`."
            )
        elif resolved_year is None and workbook_year is not None:
            resolved_year = workbook_year
            source_parts.append("workbook sample")

    unique_source_parts = list(dict.fromkeys(source_parts))
    note_parts.extend(conflict_notes)
    if conflict_notes:
        status = "conflict"
        confidence = "low"
    elif resolved_month and resolved_year is not None:
        status = "resolved"
        confidence = "high"
    elif resolved_month or resolved_year is not None:
        status = "partial"
        confidence = "medium"
    else:
        status = "unresolved"
        confidence = "none"

    return {
        "role": role_label,
        "filename": uploaded_file.name,
        "month": resolved_month,
        "year": resolved_year,
        "source": " + ".join(unique_source_parts) if unique_source_parts else "manual confirmation",
        "confidence": confidence,
        "status": status,
        "note": " ".join(part for part in note_parts if part),
    }


def _build_upload_detection_overview(energy_files, csi_file, mes_file):
    file_detections = [
        _resolve_uploaded_file_detection(uploaded_file, role)
        for role, uploaded_file in [
            *[("Energy", file) for file in energy_files],
            ("CSI", csi_file),
            ("MES", mes_file),
        ]
        if uploaded_file is not None
    ]

    detected_months = sorted({item["month"] for item in file_detections if item["month"]})
    detected_years = sorted({item["year"] for item in file_detections if item["year"] is not None})
    sources = sorted({item["source"] for item in file_detections if item["source"] and item["status"] != "unresolved"})

    blocking_issues = []
    for item in file_detections:
        if item["status"] == "conflict":
            blocking_issues.append(f"{item['role']} file `{item['filename']}` has conflicting month/year signals.")
    if len(detected_months) > 1:
        blocking_issues.append(
            "Uploaded files point to multiple months: " + ", ".join(detected_months) + "."
        )
    if len(detected_years) > 1:
        blocking_issues.append(
            "Uploaded files point to multiple years: " + ", ".join(str(year) for year in detected_years) + "."
        )

    if file_detections:
        confidence_rank = {"none": 0, "low": 1, "medium": 2, "high": 3}
        overview_confidence = max(file_detections, key=lambda item: confidence_rank[item["confidence"]])[
            "confidence"
        ]
    else:
        overview_confidence = "none"

    return {
        "signature": _build_upload_signature(energy_files, csi_file, mes_file),
        "file_detections": file_detections,
        "detected_month": detected_months[0] if len(detected_months) == 1 else None,
        "detected_year": detected_years[0] if len(detected_years) == 1 else None,
        "blocking_issues": blocking_issues,
        "source_summary": " + ".join(sources) if sources else "manual confirmation",
        "confidence": overview_confidence,
        "has_files": bool(file_detections),
    }


def _build_year_options(current_year, detected_year=None):
    year_options = list(range(current_year - 5, current_year + 2))
    if detected_year is not None and detected_year not in year_options:
        year_options.append(detected_year)
        year_options.sort()
    return year_options


def _sync_upload_target_state(detection_overview):
    current_year = datetime.now().year
    if "etl_selected_month" not in st.session_state:
        st.session_state.etl_selected_month = "June"
    if "etl_selected_year" not in st.session_state:
        st.session_state.etl_selected_year = current_year
    if "etl_manual_override" not in st.session_state:
        st.session_state.etl_manual_override = False

    if st.session_state.get("etl_upload_signature") == detection_overview["signature"]:
        return

    detected_month = detection_overview.get("detected_month")
    detected_year = detection_overview.get("detected_year")

    if detected_month:
        st.session_state.etl_selected_month = detected_month
    else:
        st.session_state.etl_selected_month = st.session_state.get("etl_selected_month", "June")

    if detected_year is not None:
        st.session_state.etl_selected_year = detected_year
    else:
        st.session_state.etl_selected_year = st.session_state.get("etl_selected_year", current_year)

    st.session_state.etl_manual_override = not (detected_month or detected_year is not None)
    st.session_state.etl_upload_signature = detection_overview["signature"]


def _safe_load_run_details(details_value):
    if details_value is None:
        return {}
    try:
        return json.loads(details_value)
    except (TypeError, json.JSONDecodeError):
        return {}


def _query_month_three_way_delta(conn, current_month, previous_month):
    gained_query = """
        SELECT DISTINCT machine_id
        FROM machine_monthly_presence
        WHERE month_year = ?
          AND system_type = 'Energy'
          AND is_three_way_match = 1
          AND machine_id NOT IN (
              SELECT machine_id
              FROM machine_monthly_presence
              WHERE month_year = ?
                AND system_type = 'Energy'
                AND is_three_way_match = 1
          )
        ORDER BY machine_id
    """
    lost_query = """
        SELECT DISTINCT machine_id
        FROM machine_monthly_presence
        WHERE month_year = ?
          AND system_type = 'Energy'
          AND is_three_way_match = 1
          AND machine_id NOT IN (
              SELECT machine_id
              FROM machine_monthly_presence
              WHERE month_year = ?
                AND system_type = 'Energy'
                AND is_three_way_match = 1
          )
        ORDER BY machine_id
    """
    gained_df = pd.read_sql_query(gained_query, conn, params=(current_month, previous_month))
    lost_df = pd.read_sql_query(lost_query, conn, params=(previous_month, current_month))
    return gained_df, lost_df


def _build_latest_run_snapshot(etl_module):
    conn = sqlite3.connect(etl_module.db_path)
    try:
        latest_run_df = pd.read_sql_query(
            """
            SELECT id, month_processed, run_date, energy_files_count, three_way_matches, match_rate, details
            FROM etl_runs
            ORDER BY run_date DESC
            LIMIT 1
            """,
            conn,
        )
        if latest_run_df.empty:
            return None

        latest_run = latest_run_df.iloc[0].to_dict()
        latest_run["details"] = _safe_load_run_details(latest_run.get("details"))

        latest_month = latest_run["month_processed"]
        latest_run["run_count_for_month"] = int(
            pd.read_sql_query(
                "SELECT COUNT(*) AS row_count FROM etl_runs WHERE month_processed = ?",
                conn,
                params=(latest_month,),
            ).iloc[0]["row_count"]
        )

        month_history_df = pd.read_sql_query(
            """
            SELECT month_processed, MAX(run_date) AS latest_run_date, COUNT(*) AS run_count
            FROM etl_runs
            GROUP BY month_processed
            ORDER BY latest_run_date DESC
            """,
            conn,
        )
        latest_run["previous_distinct_month"] = (
            month_history_df.iloc[1]["month_processed"] if len(month_history_df) > 1 else None
        )

        if latest_run["previous_distinct_month"]:
            gained_df, lost_df = _query_month_three_way_delta(
                conn,
                latest_month,
                latest_run["previous_distinct_month"],
            )
        else:
            gained_df = pd.DataFrame(columns=["machine_id"])
            lost_df = pd.DataFrame(columns=["machine_id"])

        cumulative_inventory_df = etl_module.get_machine_inventory_summary(active_only=False)
        months_processed = month_history_df.sort_values("latest_run_date")["month_processed"].tolist()

        historical_gap_df = pd.DataFrame(columns=["machine_id", "appeared_in_months"])
        if latest_month:
            historical_gap_df = pd.read_sql_query(
                """
                SELECT DISTINCT m.machine_id,
                       GROUP_CONCAT(m.month_year, ', ') AS appeared_in_months
                FROM machine_monthly_presence m
                WHERE m.is_three_way_match = 1
                  AND m.system_type = 'Energy'
                  AND m.machine_id NOT IN (
                      SELECT machine_id
                      FROM machine_monthly_presence
                      WHERE month_year = ?
                        AND is_three_way_match = 1
                        AND system_type = 'Energy'
                  )
                GROUP BY m.machine_id
                ORDER BY m.machine_id
                """,
                conn,
                params=(latest_month,),
            )

        all_time_match_count = int(
            pd.read_sql_query("SELECT COUNT(*) AS row_count FROM three_way_matches", conn).iloc[0]["row_count"]
        )

        return {
            "latest_run": latest_run,
            "gained_df": gained_df,
            "lost_df": lost_df,
            "cumulative_inventory_df": cumulative_inventory_df,
            "months_processed": months_processed,
            "historical_gap_df": historical_gap_df,
            "all_time_match_count": all_time_match_count,
        }
    finally:
        conn.close()


def render_etl_page(runtime_mode: str = "standard"):
    """Main function to render the ETL Pipeline page"""
    st.header("🔄 ETL Pipeline - Monthly Data Upload")
    read_only_runtime = suppress_write_controls(runtime_mode)
    if read_only_runtime:
        info_prefix = (
            "Demo read-only mode is active."
            if normalize_runtime_mode(runtime_mode) == "demo_readonly"
            else "Pilot review mode is active."
        )
        st.info(
            f"{info_prefix} Upload/process/backfill controls and historical-run mutations are hidden. "
            "Latest run analytics and historical provenance remain available."
        )
    
    # Initialize ETL module
    etl_module = ETLPipelineModule(initialize_schema=not read_only_runtime)
    
    # Create tabs
    tab1, tab2, tab3 = st.tabs(["📤 Upload New Data", "🧭 Latest Run Snapshot", "📈 Historical Runs"])
    
    with tab1:
        _render_extension_source_availability()
        if read_only_runtime:
            _render_demo_readonly_upload_gate()
        else:
            render_upload_section(etl_module)
    
    with tab2:
        render_current_status(etl_module)
        
    with tab3:
        render_historical_runs(etl_module, read_only=read_only_runtime)


def _render_demo_readonly_upload_gate() -> None:
    st.warning(
        "Demo read-only mode hides ETL upload, processing, backfill, and month-write controls. "
        "Use `standard` mode for operational ETL work."
    )
    st.caption(
        "The reviewer/demo shell still keeps `Latest Run Snapshot` and `Historical Runs` available as read-only evidence."
    )


def _render_extension_source_availability() -> None:
    availability_df = _build_extension_source_availability_dataframe()
    if availability_df.empty:
        return

    st.markdown("#### 🗂️ Jul 2025 to Mar 2026 Historical Source Availability")
    readiness_counts = availability_df["Backfill Readiness"].value_counts()
    summary_cols = st.columns(3)
    with summary_cols[0]:
        st.metric("Ready Months", int(readiness_counts.get("Ready", 0)))
    with summary_cols[1]:
        st.metric("Ready with Flags", int(readiness_counts.get("Ready with Flags", 0)))
    with summary_cols[2]:
        st.metric("Blocked Months", int(readiness_counts.get("Blocked", 0)))

    st.caption(
        "Task13 source onboarding distinguishes month-level readiness explicitly. "
        "Partial months remain backfillable only with the documented flags and exclusions, while blocked months stay out of scope."
    )
    st.dataframe(availability_df, use_container_width=True, hide_index=True)


def render_upload_section(etl_module):
    """Render the file upload section"""
    # Check if we have processed results in session state
    if 'etl_results' in st.session_state and st.session_state.etl_results is not None:
        # Display the results instead of upload interface
        display_processing_results(st.session_state.etl_results['mapping_results'], 
                                 st.session_state.etl_results['etl'],
                                 st.session_state.etl_results.get('canonical_materialization'))
        generate_download_options(st.session_state.etl_results['mapping_results'], 
                                st.session_state.etl_results['etl'], 
                                st.session_state.etl_results['month_name'])
        
        # Add a button to clear results and return to upload
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("---")
        if st.button("🔄 Process New Month", type="primary"):
            st.session_state.etl_results = None
            st.rerun()
        return
    
    st.markdown(
        """
        ### Upload Monthly Manufacturing Data
        Upload the required files first. The page will auto-detect the target month from filenames when it can,
        keep manual override visible, and block processing if the uploaded files point to conflicting months/years.
        """
    )
    
    # File upload sections
    st.markdown("#### 1️⃣ Energy Consumption Files")
    energy_files = st.file_uploader(
        "Upload Energy files (能耗、費用報表)",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        key="energy_uploader",
        help="You can upload multiple energy files for different date ranges within the month"
    )
    
    st.markdown("#### 2️⃣ CSI Production File")
    csi_file = st.file_uploader(
        "Upload CSI file (CSI印刷心電圖報表)",
        type=['xlsx', 'xls'],
        accept_multiple_files=False,
        key="csi_uploader",
        help="Upload the monthly CSI production report"
    )
    
    st.markdown("#### 3️⃣ MES Planning File")
    mes_file = st.file_uploader(
        "Upload MES file (MES生產數據)",
        type=['xlsx', 'xls'],
        accept_multiple_files=False,
        key="mes_uploader",
        help="Upload the monthly MES planning data"
    )

    detection_overview = _build_upload_detection_overview(energy_files, csi_file, mes_file)
    _sync_upload_target_state(detection_overview)

    st.markdown("#### 🎯 Target Month Confirmation")
    st.caption(
        "One ETL run writes one month only. Multiple energy files are accepted only when they resolve to the same target month."
    )

    if detection_overview["has_files"]:
        detection_cols = st.columns(4)
        with detection_cols[0]:
            st.metric("Detected Month", detection_overview["detected_month"] or "Needs review")
        with detection_cols[1]:
            detected_year_label = (
                str(detection_overview["detected_year"])
                if detection_overview["detected_year"] is not None
                else "Needs review"
            )
            st.metric("Detected Year", detected_year_label)
        with detection_cols[2]:
            st.metric("Detection Source", detection_overview["source_summary"])
        with detection_cols[3]:
            st.metric("Confidence", detection_overview["confidence"].title())

        with st.expander("Per-file detection details", expanded=False):
            detection_rows = []
            for item in detection_overview["file_detections"]:
                month_year_label = " / ".join(
                    part
                    for part in [
                        item["month"] or "month unresolved",
                        str(item["year"]) if item["year"] is not None else "year unresolved",
                    ]
                )
                detection_rows.append(
                    {
                        "File Role": item["role"],
                        "Filename": item["filename"],
                        "Detected Target": month_year_label,
                        "Source": item["source"],
                        "Confidence": item["confidence"].title(),
                        "Notes": item["note"],
                    }
                )
            st.dataframe(pd.DataFrame(detection_rows), use_container_width=True, hide_index=True)
    else:
        st.info(
            "Upload one or more files and the page will try to prefill the target month/year from the filenames before you process."
        )

    manual_override = st.checkbox(
        "Use manual override for the target month/year",
        key="etl_manual_override",
        help="Keep this off to use the detected month/year when the uploaded files agree. Turn it on only if you need to correct the target period.",
    )

    current_year = datetime.now().year
    year_options = _build_year_options(current_year, detection_overview["detected_year"])
    month_disabled = bool(detection_overview["detected_month"]) and not manual_override
    year_disabled = detection_overview["detected_year"] is not None and not manual_override

    selection_col1, selection_col2 = st.columns(2)
    with selection_col1:
        month_name = st.selectbox(
            "Final month to write",
            MONTH_NAME_OPTIONS,
            key="etl_selected_month",
            disabled=month_disabled,
        )
    with selection_col2:
        year = st.selectbox(
            "Final year to write",
            year_options,
            key="etl_selected_year",
            disabled=year_disabled,
        )

    month_year = f"{month_name} {year}"

    if detection_overview["blocking_issues"]:
        st.warning("Resolve the uploaded-file month/year conflicts before processing:")
        for issue in detection_overview["blocking_issues"]:
            st.write(f"- {issue}")
    elif detection_overview["has_files"]:
        st.success(f"Final confirmed month that will be written: {month_year}")
        if manual_override:
            st.caption("Manual override is active. The final write month comes from your selection above.")
        elif detection_overview["detected_year"] is None:
            st.caption(
                "Month came from file detection, but the year still requires explicit confirmation because it was not reliably present in the uploaded filenames/workbook sample."
            )
        else:
            st.caption("Processing will use the detected target month/year shown above.")
    
    # Validation and processing
    if energy_files and csi_file and mes_file:
        if detection_overview["blocking_issues"]:
            st.error("⚠️ All files are uploaded, but processing is blocked until the month/year conflicts are resolved.")
        else:
            st.success(f"✅ All files uploaded for {month_year}")
        
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            if st.button(
                f"🚀 Process {month_year}",
                type="primary",
                disabled=bool(detection_overview["blocking_issues"]),
            ):
                process_uploaded_files(energy_files, csi_file, mes_file, month_year, etl_module)
    else:
        missing = []
        if not energy_files:
            missing.append("Energy files")
        if not csi_file:
            missing.append("CSI file")
        if not mes_file:
            missing.append("MES file")
        
        if missing:
            st.warning(f"⚠️ Missing: {', '.join(missing)}")


def process_uploaded_files(energy_files, csi_file, mes_file, month_year, etl_module):
    """Process the uploaded files using ETL pipeline"""
    try:
        with st.spinner(f"Processing {month_year} data..."):
            # Save uploaded files to persistent data directory for unified view
            data_dir = get_data_dir()
            data_dir.mkdir(parents=True, exist_ok=True)
            
            # Clean month name for file naming
            month_prefix = month_year.replace(' ', '_')
            
            # Save energy files with month-specific names
            energy_paths = []
            for i, file in enumerate(energy_files):
                suffix = _uploaded_file_suffix(file)
                path = data_dir / f"{month_prefix}_energy_{i+1}{suffix}"
                with open(path, "wb") as f:
                    f.write(file.getbuffer())
                energy_paths.append(str(path))
            
            # Save CSI file with month-specific name
            csi_path = data_dir / f"{month_prefix}_csi{_uploaded_file_suffix(csi_file)}"
            with open(csi_path, "wb") as f:
                f.write(csi_file.getbuffer())
            
            # Save MES file with month-specific name
            mes_path = data_dir / f"{month_prefix}_mes{_uploaded_file_suffix(mes_file)}"
            with open(mes_path, "wb") as f:
                f.write(mes_file.getbuffer())
            
            # Initialize ETL pipeline
            etl = EnhancedSmartManufacturingETL()
            
            # Process files
            st.info("Extracting data from files...")
            etl.extract_all_sources(energy_paths, str(csi_path), str(mes_path))
            _scope_etl_state_to_month(etl, month_year)
            
            st.info("Creating machine mappings...")
            mapping_results = etl.create_comprehensive_mapping()
            
            # Save results to database including actual data
            etl_module.save_etl_results(mapping_results, month_year, etl)
            
            # Clean up temp files
            for path in energy_paths + [csi_path, mes_path]:
                os.remove(path)
            
            # Auto-trigger Maintenance Integration if file exists
            st.info("🔧 Checking for maintenance data...")
            
            # Try multiple file name patterns
            maintenance_patterns = [
                f"Maintenance Record{month_year.split()[0][:3]} to {month_year.split()[0][:3]}.xlsx",
                f"maintenance_{month_year.lower().replace(' ', '_')}.xlsx",
                f"Maintenance RecordJan to Jul.xlsx",  # Known file format
                "maintenance_data.xlsx"
            ]
            
            maintenance_file = None
            repo_root = get_repo_root()
            for pattern in maintenance_patterns:
                repo_path = repo_root / pattern
                if repo_path.exists():
                    maintenance_file = str(repo_path)
                    break
                # Also check in data directory
                data_path = data_dir / pattern
                if data_path.exists():
                    maintenance_file = str(data_path)
                    break
            
            if maintenance_file:
                st.info(f"Found maintenance file: {maintenance_file}")
                try:
                    from core.maintenance_integration import integrate_maintenance_with_etl
                    
                    with st.spinner("Integrating maintenance data..."):
                        maint_result = integrate_maintenance_with_etl(
                            maintenance_file,
                            month_year,
                            db_path=etl_module.db_path,
                        )
                        
                        if maint_result:
                            # Count matched records
                            maintenance_df = maint_result['maintenance_records']
                            matched = len(maintenance_df[maintenance_df['is_three_way_match'] == 1])
                            total = len(maintenance_df)
                            
                            # Save to database
                            conn = sqlite3.connect(etl_module.db_path)
                            maintenance_df.to_sql('maintenance_records', conn, if_exists='append', index=False)
                            
                            if maint_result['metrics'] is not None:
                                maint_result['metrics'].to_sql('maintenance_summary', conn, if_exists='replace')
                            
                            if maint_result['predictions'] is not None:
                                maint_result['predictions'].to_sql('maintenance_ml_features', conn, if_exists='replace', index=False)
                                
                                # Show high-risk machines
                                high_risk = maint_result['predictions'][
                                    maint_result['predictions']['risk_level'] == 'HIGH'
                                ]
                                
                                if not high_risk.empty:
                                    st.warning(f"""
                                    ⚠️ **Maintenance Alert**: {len(high_risk)} machines need immediate attention:
                                    {', '.join(high_risk['machine_id'].head(5).tolist())}
                                    """)
                            
                            conn.commit()
                            conn.close()
                            
                            st.success(f"""
                            ✅ Integrated {matched}/{total} maintenance records ({matched/total*100:.1f}% match rate)
                            View details in the Maintenance module.
                            """)
                            
                except Exception as e:
                    st.warning(f"Could not integrate maintenance data: {str(e)}")
            else:
                st.info("No maintenance file found. You can upload it separately in the Maintenance module.")

            canonical_result = None
            st.info("🔄 Materializing canonical Silver + Gold...")
            try:
                from modules.unified_view_module import auto_process_after_etl
                canonical_result = auto_process_after_etl(month_year, db_path=etl_module.db_path)

                if canonical_result['status'] == 'success':
                    silver_counts = canonical_result.get('silver_rows_materialized_by_table', {})
                    gold_rows = canonical_result.get('gold_fact_machine_hour_rows_created', 0)
                    st.success(
                        "✅ Canonical materialization complete: "
                        f"target month {canonical_result.get('target_month', month_year)} | "
                        f"Silver yes | Gold yes | fact_machine_hour rows {gold_rows:,}"
                    )
                    st.caption(
                        "Silver rows materialized: "
                        + ", ".join(
                            f"{table}={count:,}" for table, count in silver_counts.items()
                        )
                    )
                else:
                    st.warning(
                        "⚠️ Canonical materialization issue: "
                        f"{canonical_result.get('message', 'Unknown')}"
                    )
            except Exception as e:
                canonical_result = {
                    'status': 'error',
                    'target_month': month_year,
                    'silver_materialized': False,
                    'gold_materialized': False,
                    'message': str(e),
                }
                st.warning(f"⚠️ Could not materialize canonical Silver/Gold: {str(e)}")
                st.info("The Unified View page will continue to warn until canonical Gold is materialized.")

            # Store results in session state
            st.session_state.etl_results = {
                'mapping_results': mapping_results,
                'etl': etl,
                'month_name': month_year,
                'canonical_materialization': canonical_result,
            }
            
            # Trigger a rerun to display results
            st.success(f"✅ Successfully processed {month_year} data!")
            st.rerun()
            
    except Exception as e:
        st.error(f"❌ Error processing files: {str(e)}")
        st.exception(e)


def display_processing_results(mapping_results, etl, canonical_materialization=None):
    """Display the processing results"""
    st.markdown("## 📊 Processing Results")
    st.markdown("---")

    if canonical_materialization:
        st.markdown("### 🏗️ Canonical Materialization")
        if canonical_materialization.get('status') == 'success':
            silver_counts = canonical_materialization.get('silver_rows_materialized_by_table', {})
            gold_rows = canonical_materialization.get('gold_fact_machine_hour_rows_created', 0)
            st.success(
                f"Canonical Silver + Gold materialized for {canonical_materialization.get('target_month', 'selected month')}."
            )
            status_col1, status_col2, status_col3 = st.columns(3)
            with status_col1:
                st.metric("Canonical Silver", "Yes")
            with status_col2:
                st.metric("Canonical Gold", "Yes")
            with status_col3:
                st.metric("Gold Rows", f"{gold_rows:,}")
            if silver_counts:
                st.caption(
                    "Silver rows materialized: "
                    + ", ".join(f"{table}={count:,}" for table, count in silver_counts.items())
                )
        else:
            st.warning(
                "Canonical Silver/Gold materialization did not complete: "
                f"{canonical_materialization.get('message', 'Unknown issue')}"
            )
        st.markdown("<br>", unsafe_allow_html=True)
    
    # Summary metrics with better spacing
    stats = mapping_results['mapping_stats']
    
    # Use container for better visual grouping with proper spacing
    with st.container():
        st.markdown("### 📈 Data Volume Statistics")
        # Use 2 columns for wider display
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Energy Records", f"{stats['energy_original_rows']:,}", 
                     help="Total raw energy data records")
        with col2:
            st.metric("Energy Machines", stats['energy_unique_machines'],
                     help="Unique machines in Energy system")
        
        # Second row
        col3, col4 = st.columns(2)
        
        with col3:
            st.metric("CSI Machines", stats['csi_machines'],
                     help="Unique machines in CSI system")
        with col4:
            st.metric("MES Machines", stats['mes_machines'],
                     help="Unique machines in MES system")
    
    # Add space between sections
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Matching statistics with visual indicators
    with st.container():
        st.markdown("### 🎯 Matching Analysis")
        # First row of matching metrics - 2 columns
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Three-way Matches", stats['three_way_matches'], 
                     help="Machines found in all three systems")
        
        with col2:
            # Calculate two-way matches
            two_way = (len(etl.partial_matches.get('energy_csi_only', [])) +
                      len(etl.partial_matches.get('energy_mes_only', [])) +
                      len(etl.partial_matches.get('csi_mes_only', [])))
            st.metric("Two-way Matches", two_way,
                     help="Machines found in exactly two systems")
        
        # Second row - 2 columns
        col3, col4 = st.columns(2)
        
        with col3:
            single = (len(etl.single_system.get('energy_only', [])) +
                     len(etl.single_system.get('csi_only', [])) +
                     len(etl.single_system.get('mes_only', [])))
            st.metric("Single System Only", single,
                     help="Machines found in only one system")
        
        with col4:
            st.metric("MES Coverage", stats['mes_coverage_percent'],
                     help="Percentage of MES machines with matches")
        
        # Third row - 2 columns
        col5, col6 = st.columns(2)
        
        with col5:
            # Add completeness metric
            completeness = round(stats['three_way_matches'] / stats['mes_machines'] * 100, 1)
            st.metric("System Completeness", f"{completeness}%",
                     help="Three-way matches vs total MES machines")
        
        with col6:
            # Add total unique machines
            total_unique = stats['energy_unique_machines'] + stats['csi_machines'] + stats['mes_machines']
            st.metric("Total Machines", total_unique,
                     help="Sum of machines across all systems")
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    # Display matches in parallel layout
    st.markdown("## 🔗 Machine Matching Details")
    st.markdown("---")
    
    # Create two columns for Three-way and Partial matches
    col_three_way, col_partial = st.columns([1, 1])
    
    with col_three_way:
        st.markdown("### 🎯 Three-way Machine Matches")
        if mapping_results['three_way_matches']:
            matches_df = pd.DataFrame(mapping_results['three_way_matches'])
            display_df = matches_df[['machine_id', 'csi', 'mes', 'total_kwh']].copy()
            display_df['total_kwh'] = display_df['total_kwh'].round(2)
            
            st.dataframe(
                display_df,  # Show all matches
                use_container_width=True,
                hide_index=True,
                height=600,  # Increased height for better viewing
                column_config={
                    "machine_id": st.column_config.TextColumn("Energy ID"),
                    "csi": st.column_config.TextColumn("CSI ID"),
                    "mes": st.column_config.TextColumn("MES ID"),
                    "total_kwh": st.column_config.NumberColumn("kWh", format="%.0f")
                }
            )
            st.caption(f"Total: {len(display_df)} three-way matches")
        else:
            st.info("No three-way matches found")
    
    with col_partial:
        st.markdown("### 🔀 Partial Matches")
        
        # Create tabs for different partial match types
        tab1, tab2, tab3 = st.tabs(["Energy-CSI", "Energy-MES", "CSI-MES"])
        
        with tab1:
            if etl.partial_matches.get('energy_csi_only'):
                partial_df = pd.DataFrame(etl.partial_matches['energy_csi_only'])
                display_cols = ['energy', 'csi']
                if 'total_kwh' in partial_df.columns:
                    display_cols.append('total_kwh')
                    partial_df['total_kwh'] = partial_df['total_kwh'].round(0)
                st.dataframe(
                    partial_df[display_cols],  # Show all matches
                    hide_index=True,
                    use_container_width=True,
                    height=500  # Increased height
                )
                st.caption(f"Total: {len(partial_df)} Energy-CSI matches")
            else:
                st.info("No Energy-CSI matches")
        
        with tab2:
            if etl.partial_matches.get('energy_mes_only'):
                partial_df = pd.DataFrame(etl.partial_matches['energy_mes_only'])
                display_cols = ['energy', 'mes']
                if 'total_kwh' in partial_df.columns:
                    display_cols.append('total_kwh')
                    partial_df['total_kwh'] = partial_df['total_kwh'].round(0)
                st.dataframe(
                    partial_df[display_cols],  # Show all matches
                    hide_index=True,
                    use_container_width=True,
                    height=500  # Increased height
                )
                st.caption(f"Total: {len(partial_df)} Energy-MES matches")
            else:
                st.info("No Energy-MES matches")
        
        with tab3:
            if etl.partial_matches.get('csi_mes_only'):
                partial_df = pd.DataFrame(etl.partial_matches['csi_mes_only'])
                st.dataframe(
                    partial_df[['csi', 'mes']],  # Show all matches
                    hide_index=True,
                    use_container_width=True,
                    height=500  # Increased height
                )
                st.caption(f"Total: {len(partial_df)} CSI-MES matches")
            else:
                st.info("No CSI-MES matches")
    
    # Single System Machines section - Full width layout
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("## 🔍 Single System Machines")
    st.markdown("---")
    st.markdown("**Machines found in only one system with no matches across other systems:**")
    
    # Create tabs for better organization
    tab_energy, tab_csi, tab_mes = st.tabs(["⚡ Energy Only", "🏭 CSI Only", "📋 MES Only"])
    
    with tab_energy:
        if etl.single_system.get('energy_only'):
            energy_only = etl.single_system['energy_only']
            
            # Display total count as simple text
            st.info(f"Found {len(energy_only)} unmatched Energy machines")
            
            # Display all machines in a scrollable dataframe
            # Extract clean machine IDs if they are stored as complex objects
            clean_ids = []
            for item in energy_only:
                if isinstance(item, dict):
                    clean_ids.append(item.get('machine_id', str(item)))
                elif isinstance(item, str):
                    clean_ids.append(item)
                else:
                    clean_ids.append(str(item))
            
            energy_df = pd.DataFrame({
                'Machine ID': clean_ids,
                'System': ['Energy'] * len(clean_ids),
                'Status': ['Unmatched'] * len(clean_ids)
            })
            st.dataframe(
                energy_df,
                hide_index=True,
                use_container_width=True,
                height=400  # Fixed height with scroll
            )
            st.caption(f"Total: {len(energy_only)} unmatched Energy machines")
        else:
            st.success("✅ All Energy machines have matches!")
    
    with tab_csi:
        if etl.single_system.get('csi_only'):
            csi_only = etl.single_system['csi_only']
            
            # Display total count as simple text
            st.info(f"Found {len(csi_only)} unmatched CSI machines")
            
            # Display all machines in a scrollable dataframe
            # Extract clean machine IDs if they are stored as complex objects
            clean_ids = []
            for item in csi_only:
                if isinstance(item, dict):
                    clean_ids.append(item.get('machine_id', str(item)))
                elif isinstance(item, str):
                    clean_ids.append(item)
                else:
                    clean_ids.append(str(item))
            
            csi_df = pd.DataFrame({
                'Machine ID': clean_ids,
                'System': ['CSI'] * len(clean_ids),
                'Status': ['Unmatched'] * len(clean_ids)
            })
            st.dataframe(
                csi_df,
                hide_index=True,
                use_container_width=True,
                height=400  # Fixed height with scroll
            )
            st.caption(f"Total: {len(csi_only)} unmatched CSI machines")
        else:
            st.success("✅ All CSI machines have matches!")
    
    with tab_mes:
        if etl.single_system.get('mes_only'):
            mes_only = etl.single_system['mes_only']
            
            # Display total count as simple text
            st.info(f"Found {len(mes_only)} unmatched MES machines")
            
            # Display all machines in a scrollable dataframe
            # Extract clean machine IDs if they are stored as complex objects
            clean_ids = []
            for item in mes_only:
                if isinstance(item, dict):
                    clean_ids.append(item.get('machine_id', str(item)))
                elif isinstance(item, str):
                    clean_ids.append(item)
                else:
                    clean_ids.append(str(item))
            
            mes_df = pd.DataFrame({
                'Machine ID': clean_ids,
                'System': ['MES'] * len(clean_ids),
                'Status': ['Unmatched'] * len(clean_ids)
            })
            st.dataframe(
                mes_df,
                hide_index=True,
                use_container_width=True,
                height=400  # Fixed height with scroll
            )
            st.caption(f"Total: {len(mes_only)} unmatched MES machines")
        else:
            st.success("✅ All MES machines have matches!")


def generate_download_options(mapping_results, etl, month_name):
    """Generate downloadable reports"""
    st.markdown("### 💾 Download Results")
    
    # Use wider columns with specific widths
    col1, col2, col3 = st.columns([1.2, 1, 1])
    
    with col1:
        # Generate Excel report in memory if not already done
        excel_key = f"excel_report_{month_name}"
        if excel_key not in st.session_state:
            with st.spinner("Preparing Excel report..."):
                report_name = f"{month_name.lower()}_etl_report.xlsx"
                etl.generate_enhanced_report(report_name)
                
                # Read the file and store in session state
                with open(report_name, "rb") as f:
                    st.session_state[excel_key] = f.read()
                
                # Clean up temp file
                os.remove(report_name)
        
        # Provide download button
        st.download_button(
            label="📊 Excel Report",
            data=st.session_state[excel_key],
            file_name=f"{month_name.lower()}_etl_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="excel_download"
        )
    
    with col2:
        # Convert numpy types to Python native types for JSON serialization
        def convert_to_serializable(obj):
            """Convert numpy types to Python native types"""
            import numpy as np
            if isinstance(obj, dict):
                return {k: convert_to_serializable(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_to_serializable(item) for item in obj]
            elif isinstance(obj, (np.integer, np.int64)):
                return int(obj)
            elif isinstance(obj, (np.floating, np.float64)):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif pd.isna(obj):
                return None
            else:
                return obj
        
        # Convert mapping results to be JSON serializable
        serializable_results = convert_to_serializable(mapping_results)
        json_data = json.dumps(serializable_results, indent=2, ensure_ascii=False)
        
        st.download_button(
            label="📥 JSON Data",
            data=json_data,
            file_name=f"{month_name.lower()}_mappings.json",
            mime="application/json"
        )
    
    with col3:
        # Download integrated metrics
        if hasattr(etl, 'integrated_metrics') and len(etl.integrated_metrics) > 0:
            csv_data = etl.integrated_metrics.to_csv(index=False)
            st.download_button(
                label="📥 CSV Metrics",
                data=csv_data,
                file_name=f"{month_name.lower()}_integrated_metrics.csv",
                mime="text/csv"
            )


def render_current_status(etl_module):
    """Render the latest ETL snapshot with action-oriented context."""
    st.markdown("### 🧭 Latest Run Snapshot")
    st.caption("Answering three questions: what was processed last, what changed, and what can the operator do next.")

    snapshot = _build_latest_run_snapshot(etl_module)
    if snapshot is None:
        st.info("No data processed yet. Please upload files to begin.")
        return

    latest_run = snapshot["latest_run"]
    details = latest_run["details"]
    latest_month = latest_run["month_processed"]
    previous_month = latest_run["previous_distinct_month"]
    gained_df = snapshot["gained_df"]
    lost_df = snapshot["lost_df"]

    st.info(
        f"Latest processed month: **{latest_month}** | "
        f"Last run recorded: **{pd.to_datetime(latest_run['run_date']).strftime('%Y-%m-%d %H:%M')}** | "
        f"Run records retained for this month: **{latest_run['run_count_for_month']}**"
    )

    metric_col1, metric_col2, metric_col3, metric_col4, metric_col5 = st.columns(5)
    with metric_col1:
        st.metric("Three-way Matches", latest_run["three_way_matches"])
    with metric_col2:
        st.metric("Match Rate", f"{float(latest_run['match_rate']):.1f}%")
    with metric_col3:
        st.metric("Energy Machines", details.get("energy_unique_machines", 0))
    with metric_col4:
        st.metric("CSI Machines", details.get("csi_machines", 0))
    with metric_col5:
        st.metric("MES Machines", details.get("mes_machines", 0))

    st.markdown("#### 🔄 What Changed Since The Previous Distinct Month")
    if previous_month:
        delta_col1, delta_col2 = st.columns(2)
        with delta_col1:
            st.metric(
                f"Gained Three-Way Matches vs {previous_month}",
                f"{len(gained_df):,}",
            )
            if not gained_df.empty:
                with st.expander(f"View gained matches ({len(gained_df)})", expanded=False):
                    st.write(gained_df["machine_id"].tolist())
        with delta_col2:
            st.metric(
                f"Lost Three-Way Matches vs {previous_month}",
                f"{len(lost_df):,}",
            )
            if not lost_df.empty:
                with st.expander(f"View lost matches ({len(lost_df)})", expanded=False):
                    st.write(lost_df["machine_id"].tolist())
        if gained_df.empty and lost_df.empty:
            st.caption(f"No three-way match movement was recorded between {previous_month} and {latest_month}.")
    else:
        st.caption("Month-over-month gained/lost counts will appear once at least two distinct processed months exist.")

    st.markdown("#### ✅ What You Can Do Next")
    next_steps = [
        f"Upload a new month when you want to add the next canonical month after {latest_month}.",
        f"Rerun {latest_month} only when corrected source files are ready; the active month snapshot will be replaced for that month.",
        "Use the Historical Runs tab for provenance, export, and run-to-run presentation comparison.",
    ]
    for step in next_steps:
        st.write(f"- {step}")

    with st.expander("Historical inventory context", expanded=False):
        inventory_df = snapshot["cumulative_inventory_df"]
        if not inventory_df.empty:
            inventory_col1, inventory_col2, inventory_col3, inventory_col4 = st.columns(4)
            with inventory_col1:
                energy_total = inventory_df.loc[
                    inventory_df["system_type"] == "Energy", "machine_count"
                ].sum()
                st.metric("Energy Machines (All Time)", int(energy_total))
            with inventory_col2:
                csi_total = inventory_df.loc[
                    inventory_df["system_type"] == "CSI", "machine_count"
                ].sum()
                st.metric("CSI Machines (All Time)", int(csi_total))
            with inventory_col3:
                mes_total = inventory_df.loc[
                    inventory_df["system_type"] == "MES", "machine_count"
                ].sum()
                st.metric("MES Machines (All Time)", int(mes_total))
            with inventory_col4:
                st.metric(
                    "Unique Three-Way Matches (All Time)",
                    f"{snapshot['all_time_match_count']:,}",
                )
            st.caption(
                "Processed months in historical order: "
                + (", ".join(snapshot["months_processed"]) if snapshot["months_processed"] else "none")
            )

        historical_gap_df = snapshot["historical_gap_df"]
        if not historical_gap_df.empty:
            st.markdown(f"Machines historically matched but not present in the latest month `{latest_month}`:")
            display_df = historical_gap_df.rename(
                columns={"machine_id": "Machine ID", "appeared_in_months": "Appeared In"}
            )
            st.dataframe(display_df, use_container_width=True, hide_index=True)
            st.download_button(
                label="📥 Download Historical Machines CSV",
                data=display_df.to_csv(index=False),
                file_name=f"historical_machines_not_in_{latest_month.replace(' ', '_')}.csv",
                mime="text/csv",
            )
        else:
            st.success(f"✅ All historically matched machines are present in {latest_month}")


def render_historical_runs(etl_module, *, read_only: bool = False):
    """Render historical ETL runs with delete and reorder functionality"""
    st.markdown("### 📈 Historical Processing Runs")
    st.caption(
        "Rerunning the same month replaces the active month snapshot in ETL storage and canonical materialization, "
        "while the run records below stay available for provenance and presentation comparison."
    )
    if read_only:
        st.info(
            "Demo read-only mode hides delete, reorder, and bulk-selection controls for ETL run history."
        )
    
    # Initialize session state for bulk selection
    if 'selected_runs' not in st.session_state:
        st.session_state.selected_runs = []
    if 'bulk_select_mode' not in st.session_state:
        st.session_state.bulk_select_mode = False
    
    # Toolbar
    col1, col2, col3, col4 = st.columns([1, 1, 1, 2])
    
    with col1:
        if (not read_only) and st.button("🔄 Bulk Select", type="secondary" if not st.session_state.bulk_select_mode else "primary"):
            st.session_state.bulk_select_mode = not st.session_state.bulk_select_mode
            st.session_state.selected_runs = []
    
    with col2:
        sort_option = st.selectbox(
            "Sort By",
            ["Custom Order", "Date (Newest)", "Date (Oldest)", "Month", "Match Rate"],
            key="sort_option"
        )
        
        # Map sort options to order_by parameter
        sort_mapping = {
            "Custom Order": "display_order",
            "Date (Newest)": "date_desc",
            "Date (Oldest)": "date_asc",
            "Month": "month",
            "Match Rate": "match_rate"
        }
        order_by = sort_mapping[sort_option]
    
    with col3:
        if (not read_only) and st.button("↺ Reset Order"):
            etl_module.reset_display_order()
            st.success("Order reset to chronological!")
            st.rerun()
    
    with col4:
        if (not read_only) and st.session_state.bulk_select_mode and len(st.session_state.selected_runs) > 0:
            if st.button(f"🗑️ Delete Selected ({len(st.session_state.selected_runs)})", type="primary"):
                st.session_state.confirm_bulk_delete = True
    
    # Bulk delete confirmation dialog
    if (not read_only) and st.session_state.get('confirm_bulk_delete', False):
        st.warning(f"⚠️ Are you sure you want to delete {len(st.session_state.selected_runs)} selected records?")
        col_yes, col_no = st.columns(2)
        with col_yes:
            if st.button("Yes, Delete All", type="primary", key="confirm_yes"):
                etl_module.delete_multiple_runs(st.session_state.selected_runs)
                st.success(f"Deleted {len(st.session_state.selected_runs)} records!")
                st.session_state.selected_runs = []
                st.session_state.bulk_select_mode = False
                st.session_state.confirm_bulk_delete = False
                st.rerun()
        with col_no:
            if st.button("Cancel", key="confirm_no"):
                st.session_state.confirm_bulk_delete = False
                st.rerun()
    
    # Get data
    history_df = etl_module.get_historical_summary(order_by=order_by)
    
    if not history_df.empty:
        # Format the dataframe
        history_df['run_date_formatted'] = pd.to_datetime(history_df['run_date']).dt.strftime('%Y-%m-%d %H:%M')
        history_df['match_rate_formatted'] = history_df['match_rate'].apply(lambda x: f"{x:.1f}%")
        
        # Display records with actions
        for idx, row in history_df.iterrows():
            col_check, col_data, col_actions = st.columns([0.5, 8, 1.5])
            
            # Checkbox for bulk selection
            with col_check:
                if st.session_state.bulk_select_mode and not read_only:
                    if st.checkbox("", key=f"select_{row['id']}", value=row['id'] in st.session_state.selected_runs):
                        if row['id'] not in st.session_state.selected_runs:
                            st.session_state.selected_runs.append(row['id'])
                    else:
                        if row['id'] in st.session_state.selected_runs:
                            st.session_state.selected_runs.remove(row['id'])
            
            # Data display
            with col_data:
                # Create a container for better formatting
                with st.container():
                    data_cols = st.columns([2, 3, 1.5, 1.5])
                    with data_cols[0]:
                        st.write(f"**{row['month_processed']}**")
                    with data_cols[1]:
                        st.write(row['run_date_formatted'])
                    with data_cols[2]:
                        st.write(f"Matches: {row['three_way_matches']}")
                    with data_cols[3]:
                        st.write(f"Rate: {row['match_rate_formatted']}")
            
            # Action buttons
            with col_actions:
                if not st.session_state.bulk_select_mode and not read_only:
                    action_cols = st.columns(3)
                    
                    # Up button
                    with action_cols[0]:
                        if st.button("↑", key=f"up_{row['id']}", help="Move up"):
                            etl_module.update_display_order(row['id'], 'up')
                            st.rerun()
                    
                    # Down button
                    with action_cols[1]:
                        if st.button("↓", key=f"down_{row['id']}", help="Move down"):
                            etl_module.update_display_order(row['id'], 'down')
                            st.rerun()
                    
                    # Delete button
                    with action_cols[2]:
                        if st.button("🗑️", key=f"del_{row['id']}", help="Delete"):
                            # Store the ID to delete in session state
                            st.session_state[f"confirm_delete_{row['id']}"] = True
                            st.rerun()
            
            # Confirmation dialog for individual deletion
            if (not read_only) and st.session_state.get(f"confirm_delete_{row['id']}", False):
                with st.container():
                    st.warning(f"Are you sure you want to delete the {row['month_processed']} run from {row['run_date_formatted']}?")
                    col_yes, col_no = st.columns(2)
                    with col_yes:
                        if st.button("Yes, Delete", key=f"yes_del_{row['id']}", type="primary"):
                            etl_module.delete_etl_run(row['id'])
                            st.success("Record deleted!")
                            del st.session_state[f"confirm_delete_{row['id']}"]
                            st.rerun()
                    with col_no:
                        if st.button("Cancel", key=f"no_del_{row['id']}"):
                            del st.session_state[f"confirm_delete_{row['id']}"]
                            st.rerun()
            
            st.divider()
        
        # Summary statistics
        st.markdown("#### 📊 Summary Statistics")
        
        # Calculate statistics
        total_runs = len(history_df)
        avg_match_rate = history_df['match_rate'].mean()
        
        stat_cols = st.columns(2)
        with stat_cols[0]:
            st.metric("Total Runs", total_runs)
        with stat_cols[1]:
            st.metric("Average Match Rate", f"{avg_match_rate:.1f}%")
        
        # Trend chart
        if len(history_df) > 1:
            st.markdown("#### 📈 Match Rate Trend")
            # Create a proper dataframe for the chart
            chart_df = history_df.copy()
            chart_df['match_rate_num'] = chart_df['match_rate']
            
            # Parse month_processed to create a proper date for x-axis
            def parse_month_year(month_str):
                """Convert 'Month Year' or 'Month' to a date object"""
                if ' ' in month_str:
                    # Already has year
                    month, year = month_str.split(' ')
                else:
                    # Just month, assume current year
                    month = month_str
                    year = str(datetime.now().year)
                
                # Convert month name to number
                month_num = datetime.strptime(month, '%B').month
                # Create date object (first day of the month)
                return datetime(int(year), month_num, 1)
            
            chart_df['month_date'] = chart_df['month_processed'].apply(parse_month_year)
            chart_df = chart_df.sort_values('month_date')  # Ensure chronological order
            
            # Use altair for better control
            import altair as alt
            
            chart = alt.Chart(chart_df).mark_line(point=True).encode(
                x=alt.X('month_date:T', 
                    title='Month',
                    axis=alt.Axis(format='%b %Y')  # Show as "Jan 2025", "Feb 2025", etc.
                ),
                y=alt.Y('match_rate_num:Q', 
                    title='Match Rate (%)', 
                    scale=alt.Scale(domain=[0, 100])
                ),
                tooltip=[
                    alt.Tooltip('month_processed:N', title='Month'),
                    alt.Tooltip('run_date_formatted:N', title='Processed On'),
                    alt.Tooltip('match_rate_formatted:N', title='Match Rate'),
                    alt.Tooltip('three_way_matches:Q', title='Three-way Matches')
                ]
            ).properties(
                height=300
            ).interactive()
            
            st.altair_chart(chart, use_container_width=True)
        
        # Export option
        st.markdown("#### 💾 Export Data")
        csv = history_df[['month_processed', 'run_date', 'three_way_matches', 'match_rate']].to_csv(index=False)
        st.download_button(
            label="📥 Download History as CSV",
            data=csv,
            file_name="etl_run_history.csv",
            mime="text/csv"
        )
        
    else:
        st.info("No historical runs yet. Process your first month to see history.")


# Module ready for import
